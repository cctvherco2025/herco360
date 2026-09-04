"""Formularios personalizados — "Haz tu form personalizado".

Cualquiera con acceso al módulo Formulario (FLOS o Rutina Operativa) puede
armar un formulario propio: título, preguntas (opción única, casillas o
texto), decide si cada pregunta lleva puntaje, y a quién va dirigido
(todos / un área / un cargo). El esquema se guarda en Mongo (no en código),
y este router sirve tanto para crearlo como para responderlo y consultarlo —
el "motor" es genérico, no conoce de antemano las preguntas.

Igual que FLOS/Rutina: la respuesta completa se envía de un solo POST al
terminar, con las fotos subidas a R2/almacenamiento (storage.py) y solo
descargables autenticado.

Visibilidad de un formulario ya creado:
  - Responderlo: admin, Director comercial, el creador, o quien calce con la
    audiencia elegida (todos / su área / su cargo) — "todos" no exige tener
    acceso al módulo: es a propósito para poder alcanzar a cualquier persona
    de la empresa con una encuesta puntual.
  - Ver todas las respuestas: admin, Director comercial o el creador.
    Cualquier otra persona solo ve las respuestas que ella misma envió.
"""
import io
import json
import logging
from typing import List, Optional

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form

from fastapi.responses import Response, StreamingResponse

from core import db, get_current_user, serialize_doc, new_id, now_iso, can_use_formulario_module, can_manage_promos, require_promo_access
from models import CustomFormInput
import storage

router = APIRouter(prefix='/formularios-custom', tags=['formularios-custom'])
logger = logging.getLogger('formularios_custom')

TIPOS_VALIDOS = {'opcion_unica', 'checklist', 'texto'}
MAX_PHOTO_SIZE = 8 * 1024 * 1024  # 8 MB por foto
MAX_ITEMS = 60
MAX_EXCEL_ROWS = 500

_MAIN_COL_KEYWORDS = ('promocion', 'promoción', 'producto', 'articulo', 'artículo', 'descripcion', 'descripción', 'nombre')


async def require_builder_access(user=Depends(get_current_user)):
    if not can_use_formulario_module(user):
        raise HTTPException(status_code=403, detail='No tienes acceso al módulo Formulario')
    return user


def _item_max(item: dict) -> int:
    if not item.get('scored'):
        return 0
    pts = [int(o.get('pts') or 0) for o in item.get('opciones', [])]
    if not pts:
        return 0
    return sum(pts) if item.get('tipo') == 'checklist' else max(pts)


def _audience_match(user, audiencia: dict) -> bool:
    """Audiencia por listas: todos, o cualquier combinación de área / cargo /
    usuario individual — basta con calzar en una de ellas."""
    audiencia = audiencia or {}
    if audiencia.get('todos'):
        return True
    if user.get('id') in (audiencia.get('user_ids') or []):
        return True
    if (user.get('area') or '').strip() in (audiencia.get('areas') or []):
        return True
    if (user.get('position') or '').strip() in (audiencia.get('cargos') or []):
        return True
    return False


async def _resolve_audience_users(audiencia: dict) -> list:
    """IDs de todos los usuarios aprobados que hoy calzan con la audiencia —
    se guarda como foto fija al publicar (audiencia_resueltos), así que si
    alguien cambia de cargo/área después, el historial de esa publicación no
    se mueve."""
    audiencia = audiencia or {}
    query = {'status': 'approved'}
    if not audiencia.get('todos'):
        ors = []
        if audiencia.get('areas'):
            ors.append({'area': {'$in': audiencia['areas']}})
        if audiencia.get('cargos'):
            ors.append({'position': {'$in': audiencia['cargos']}})
        if audiencia.get('user_ids'):
            ors.append({'id': {'$in': audiencia['user_ids']}})
        if not ors:
            return []
        query['$or'] = ors
    users = await db.users.find(query, {'_id': 0, 'id': 1}).to_list(2000)
    return [u['id'] for u in users]


def _can_fill(user, form: dict) -> bool:
    if form.get('status') == 'borrador' and form.get('creator_id') != user.get('id'):
        return False
    if user.get('role') == 'admin':
        return True
    if (user.get('position') or '').strip() == 'Director comercial':
        return True
    if form.get('creator_id') == user.get('id'):
        return True
    return _audience_match(user, form.get('audiencia'))


def _sees_all_responses(user, form: dict) -> bool:
    if user.get('role') == 'admin':
        return True
    if (user.get('position') or '').strip() == 'Director comercial':
        return True
    return form.get('creator_id') == user.get('id')


async def _get_form_or_404(form_id: str) -> dict:
    form = await db.custom_forms.find_one({'id': form_id}, {'_id': 0})
    if not form:
        raise HTTPException(status_code=404, detail='Formulario no encontrado')
    return form


def _guess_main_column(headers: list, rows: list) -> Optional[str]:
    """Heurística para adivinar cuál columna trae el nombre de la
    promoción/producto: primero por nombre de columna (palabras clave
    comunes), si ninguna calza cae a la primera columna cuyo contenido es
    mayormente texto (no numérico) en la muestra de filas."""
    norm = [h.strip().lower() for h in headers]
    for kw in _MAIN_COL_KEYWORDS:
        for i, h in enumerate(norm):
            if kw in h:
                return headers[i]
    for h in headers:
        non_empty, text_like = 0, 0
        for row in rows[:50]:
            v = row.get(h)
            if v in (None, ''):
                continue
            non_empty += 1
            if not isinstance(v, (int, float)):
                text_like += 1
        if non_empty and text_like / non_empty > 0.7:
            return h
    return headers[0] if headers else None


# --------------------------------------------------------------------------- #
#  Promociones del mes — plantilla de Excel para guiar al usuario en el paso
#  "Cargar Excel". Cualquier estructura de columnas funciona igual (Fase 1 la
#  detecta sola); esta plantilla es solo una guía con un ejemplo llenado.
# --------------------------------------------------------------------------- #
@router.get('/promociones/plantilla')
async def promo_template(user=Depends(require_promo_access)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Promociones'
    ws.append(['Promoción', 'Categoría', 'Descuento', 'Vigencia'])
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='16A34A')
        c.alignment = Alignment(horizontal='center')
    ws.append(['2x1 Detergente Ariel 900ml', 'Limpieza', '50%', '01 al 30 de septiembre'])
    ws.append(["Combo desayuno Kellogg's", 'Alimentos', '20%', '01 al 15 de septiembre'])
    ws.append(['Descuento llantas Goodyear', 'Automotriz', '15%', 'Todo el mes'])
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 26

    ws2 = wb.create_sheet('Instrucciones')
    for row in [
        ['Instrucciones para Promociones del mes'],
        [''],
        ['Columna', 'Descripción'],
        ['Promoción', 'Obligatoria. El nombre de cada promoción o producto — cada fila se convierte en una pregunta del formulario.'],
        ['Categoría / Descuento / Vigencia', 'Opcionales. Se muestran como referencia junto a cada pregunta; puedes agregar, quitar o renombrar estas columnas libremente.'],
        [''],
        ['Nota', 'La estructura de columnas no está fija: al subir tu Excel, HERCO360 detecta las columnas reales y te deja elegir cuál es el nombre de la promoción.'],
        ['Nota', 'Cada promoción se pregunta como "¿La promoción está visible en tienda?" con opciones Sí / No / No aplica, más un campo de notas para observaciones.'],
    ]:
        ws2.append(row)
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="plantilla_promociones_herco360.xlsx"'},
    )


# --------------------------------------------------------------------------- #
#  Promociones del mes — inspección de Excel (paso "Cargar Excel" / "Detectar
#  columnas"). No guarda nada: el front arma las preguntas con esta data y
#  las manda ya armadas al crear el formulario (create_form de abajo).
# --------------------------------------------------------------------------- #
@router.post('/inspeccionar-excel')
async def inspect_excel(file: UploadFile = File(...), user=Depends(require_promo_access)):
    fn = (file.filename or '').lower()
    if fn.endswith('.xls') and not fn.endswith('.xlsx'):
        raise HTTPException(status_code=400,
                             detail='Ese formato .xls antiguo no se puede leer directo — guarda el archivo como .xlsx desde Excel e inténtalo de nuevo')
    if not fn.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail='El archivo debe ser .xlsx')

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail='El archivo está vacío')

    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_raw = list(ws.iter_rows(values_only=True))
    except Exception:
        raise HTTPException(status_code=400, detail='No se pudo leer el archivo Excel')
    if not rows_raw:
        raise HTTPException(status_code=400, detail='El archivo no tiene datos')

    header_row = rows_raw[0]
    headers, seen = [], set()
    for i, c in enumerate(header_row):
        name = str(c).strip() if (c is not None and str(c).strip()) else f'Columna {i + 1}'
        base, n = name, 2
        while name in seen:  # encabezados duplicados o vacíos repetidos
            name = f'{base} ({n})'; n += 1
        seen.add(name)
        headers.append(name)

    rows_out = []
    truncated = False
    for row in rows_raw[1:]:
        if row is None or all((c is None or str(c).strip() == '') for c in row):
            continue
        if len(rows_out) >= MAX_EXCEL_ROWS:
            truncated = True
            break
        item = {}
        for i, h in enumerate(headers):
            v = row[i] if i < len(row) else None
            if v is not None and not isinstance(v, (int, float)):
                v = str(v).strip()
            item[h] = v
        rows_out.append(item)

    if not rows_out:
        raise HTTPException(status_code=400, detail='No se encontraron filas con datos debajo del encabezado')

    return {
        'headers': headers,
        'suggested_main_column': _guess_main_column(headers, rows_out),
        'total_rows': len(rows_out),
        'truncated': truncated,
        'rows': rows_out,
    }


# --------------------------------------------------------------------------- #
#  Construir / administrar formularios
# --------------------------------------------------------------------------- #
@router.post('')
async def create_form(data: CustomFormInput, user=Depends(require_builder_access)):
    kind = (data.kind or 'generic').strip()
    if kind not in ('generic', 'promociones'):
        raise HTTPException(status_code=400, detail='Tipo de formulario inválido')
    if kind == 'promociones' and not can_manage_promos(user):
        raise HTTPException(status_code=403, detail='No tienes permiso para publicar Promociones del mes')
    status = (data.status or 'publicado').strip()
    if status not in ('borrador', 'publicado'):
        raise HTTPException(status_code=400, detail='Estado inválido')

    titulo = data.titulo.strip()
    if not titulo:
        raise HTTPException(status_code=400, detail='Indica un título para el formulario')
    aud = data.audiencia
    if not (aud.todos or aud.areas or aud.cargos or aud.user_ids):
        raise HTTPException(status_code=400, detail='Indica a quién va dirigido el formulario')
    if not data.items:
        raise HTTPException(status_code=400, detail='Agrega al menos una pregunta')
    if len(data.items) > MAX_ITEMS:
        raise HTTPException(status_code=400, detail=f'Máximo {MAX_ITEMS} preguntas por formulario')

    items = []
    for it in data.items:
        tipo = it.tipo.strip()
        if tipo not in TIPOS_VALIDOS:
            raise HTTPException(status_code=400, detail=f'Tipo de pregunta inválido: {tipo}')
        titulo_it = it.titulo.strip()
        if not titulo_it:
            raise HTTPException(status_code=400, detail='Cada pregunta necesita un título')
        scored = bool(it.scored) and tipo != 'texto'  # el texto libre nunca puntúa
        opciones = []
        if tipo in ('opcion_unica', 'checklist'):
            for o in it.opciones:
                label = (o.label or '').strip()
                if not label:
                    continue
                opciones.append({'label': label, 'pts': int(o.pts) if (scored and o.pts is not None) else None})
            if not opciones:
                raise HTTPException(status_code=400, detail=f'"{titulo_it}" necesita al menos una opción')
            if scored and any(o['pts'] is None for o in opciones):
                raise HTTPException(status_code=400, detail=f'"{titulo_it}": asigna puntos a todas sus opciones')
        item = {
            'id': it.id or new_id(),
            'seccion': (it.seccion or 'General').strip() or 'General',
            'titulo': titulo_it,
            'pregunta': (it.pregunta or '').strip(),
            'tipo': tipo,
            'scored': scored,
            'opciones': opciones,
            'permite_foto': bool(it.permite_foto),
        }
        item['max'] = _item_max(item)
        items.append(item)

    periodo = (data.periodo or '').strip() or None
    serie_key = (data.serie_key or '').strip() or None
    audiencia_dict = data.audiencia.model_dump()
    total_max = sum(it['max'] for it in items)
    doc = {
        'id': new_id(),
        'titulo': titulo,
        'descripcion': (data.descripcion or '').strip(),
        'kind': kind,
        'periodo': periodo,
        'serie_key': serie_key,
        'status': status,
        'creator_id': user['id'], 'creator_name': user['name'], 'creator_avatar': user.get('avatar_url'),
        'audiencia': audiencia_dict,
        'audiencia_resueltos': await _resolve_audience_users(audiencia_dict),
        'items': items,
        'has_scoring': total_max > 0,
        'total_max': total_max,
        'created_at': now_iso(),
    }
    await db.custom_forms.insert_one(doc)
    doc.pop('_id', None)
    return serialize_doc(doc)


@router.get('/disponibles')
async def list_available(kind: Optional[str] = None, user=Depends(get_current_user)):
    """Formularios que el usuario puede responder (audiencia) o que él mismo creó.

    Por defecto solo trae los 'generic' (form builder libre) — los 'promociones'
    tienen su propia tarjeta fija en el Hub, no la lista de "Formularios
    personalizados". Pasa kind=promociones para pedir justo esos."""
    forms = await db.custom_forms.find({}, {'_id': 0}).sort('created_at', -1).to_list(500)
    out = []
    for f in forms:
        f_kind = f.get('kind') or 'generic'
        if kind:
            if f_kind != kind:
                continue
        elif f_kind != 'generic':
            continue
        is_creator = f.get('creator_id') == user['id']
        if f.get('status') == 'borrador' and not is_creator:
            continue
        if not (is_creator or _can_fill(user, f)):
            continue
        out.append({
            'id': f['id'], 'titulo': f['titulo'], 'descripcion': f.get('descripcion', ''),
            'kind': f_kind, 'periodo': f.get('periodo'), 'serie_key': f.get('serie_key'),
            'status': f.get('status') or 'publicado',
            'has_scoring': f.get('has_scoring', False), 'is_creator': is_creator,
            'creator_name': f.get('creator_name'), 'audiencia': f.get('audiencia'),
            'created_at': f.get('created_at'),
        })
    return serialize_doc(out)


@router.get('/{form_id}')
async def get_form(form_id: str, user=Depends(get_current_user)):
    form = await _get_form_or_404(form_id)
    if not _can_fill(user, form):
        raise HTTPException(status_code=403, detail='No tienes acceso a este formulario')
    return serialize_doc(form)


@router.delete('/{form_id}')
async def delete_form(form_id: str, user=Depends(get_current_user)):
    form = await _get_form_or_404(form_id)
    if not (user.get('role') == 'admin' or form.get('creator_id') == user['id']):
        raise HTTPException(status_code=403, detail='Solo quien lo creó (o un admin) puede eliminarlo')
    await db.custom_form_responses.delete_many({'form_id': form_id})
    await db.custom_forms.delete_one({'id': form_id})
    return {'message': 'Formulario eliminado'}


# --------------------------------------------------------------------------- #
#  Responder
# --------------------------------------------------------------------------- #
@router.post('/{form_id}/respuestas')
async def submit_response(
    form_id: str,
    data: str = Form(...),
    photos: List[UploadFile] = File(default=[]),
    photo_owner: List[str] = Form(default=[]),
    user=Depends(get_current_user),
):
    form = await _get_form_or_404(form_id)
    if not _can_fill(user, form):
        raise HTTPException(status_code=403, detail='No tienes acceso a este formulario')

    try:
        payload = json.loads(data)
    except Exception:
        raise HTTPException(status_code=400, detail='Datos de la respuesta inválidos')
    entries_in = payload.get('entries') or []
    if not entries_in:
        raise HTTPException(status_code=400, detail='La respuesta no tiene preguntas contestadas')

    items_by_id = {it['id']: it for it in form['items']}
    clean_entries = []
    total_score, total_max = 0, 0
    for e in entries_in:
        iid = str(e.get('id') or '').strip()
        item = items_by_id.get(iid)
        if not item:
            continue
        score = 0
        if item['scored']:
            score = max(0, min(int(e.get('score') or 0), item['max']))
            total_score += score
            total_max += item['max']
        clean_entries.append({
            'id': iid, 'titulo': item['titulo'], 'seccion': item['seccion'], 'tipo': item['tipo'],
            'respuesta': e.get('respuesta') or e.get('opcion') or [],
            'score': score, 'max': item['max'], 'note': (e.get('note') or '').strip(),
            'photos': [],
        })

    if len(photos) != len(photo_owner):
        raise HTTPException(status_code=400, detail='Las fotos no coinciden con sus preguntas')

    resp_id = new_id()
    entries_by_id = {e['id']: e for e in clean_entries}
    for f, owner in zip(photos, photo_owner):
        content = await f.read()
        if not content:
            continue
        if len(content) > MAX_PHOTO_SIZE:
            raise HTTPException(status_code=400,
                                detail=f'Una foto supera el máximo de {MAX_PHOTO_SIZE // (1024 * 1024)} MB')
        ctype = f.content_type or 'image/jpeg'
        photo_id = new_id()
        ext = 'png' if 'png' in ctype else ('webp' if 'webp' in ctype else 'jpg')
        path = f'{storage.APP_NAME}/formularios-custom/{form_id}/{resp_id}/{photo_id}.{ext}'
        try:
            await storage.put_object(path, content, ctype)
        except Exception as ex:
            logger.error(f'photo upload failed: {ex}')
            raise HTTPException(status_code=502, detail='No se pudo subir una de las fotos')
        if owner in entries_by_id:
            entries_by_id[owner]['photos'].append({'id': photo_id, 'path': path, 'content_type': ctype})

    pct = round(total_score / total_max * 100) if total_max else None
    doc = {
        'id': resp_id, 'form_id': form_id, 'form_titulo': form['titulo'],
        'form_kind': form.get('kind') or 'generic', 'periodo': form.get('periodo'),
        'respondent_id': user['id'], 'respondent_name': user['name'], 'respondent_avatar': user.get('avatar_url'),
        'respondent_sucursal': user.get('sucursal') or '', 'respondent_position': user.get('position') or '',
        'entries': clean_entries,
        'total_score': total_score, 'total_max': total_max, 'percent': pct,
        'created_at': now_iso(),
    }
    await db.custom_form_responses.insert_one(doc)
    doc.pop('_id', None)
    return serialize_doc(doc)


@router.get('/{form_id}/respuestas')
async def list_responses(form_id: str, user=Depends(get_current_user)):
    form = await _get_form_or_404(form_id)
    if not (_sees_all_responses(user, form) or _can_fill(user, form)):
        raise HTTPException(status_code=403, detail='No tienes acceso a este formulario')
    q = {'form_id': form_id}
    if not _sees_all_responses(user, form):
        q['respondent_id'] = user['id']
    rows = await db.custom_form_responses.find(q, {'_id': 0, 'entries': 0}).sort('created_at', -1).limit(300).to_list(300)
    return serialize_doc(rows)


@router.get('/{form_id}/reporte')
async def get_report(form_id: str, user=Depends(get_current_user)):
    """Panel de resultados y seguimiento — pensado para "Promociones del mes"
    pero funciona para cualquier formulario con audiencia_resueltos: cruza
    los usuarios asignados (foto fija al publicar) con sus respuestas para
    armar KPIs, un reporte por pregunta/promoción y otro por sucursal.
    Mismo permiso que ver todas las respuestas (admin, Director comercial o
    el creador del formulario)."""
    form = await _get_form_or_404(form_id)
    if not _sees_all_responses(user, form):
        raise HTTPException(status_code=403, detail='No tienes acceso al reporte de este formulario')

    assigned_ids = list(dict.fromkeys(form.get('audiencia_resueltos') or []))
    assigned_users = []
    if assigned_ids:
        assigned_users = await db.users.find(
            {'id': {'$in': assigned_ids}}, {'_id': 0, 'id': 1, 'name': 1, 'sucursal': 1, 'position': 1}
        ).to_list(2000)
    users_by_id = {u['id']: u for u in assigned_users}

    responses = await db.custom_form_responses.find({'form_id': form_id}, {'_id': 0}).to_list(2000)
    # si un usuario respondió más de una vez, cuenta como "respondió" una sola vez
    responded_ids = {r['respondent_id'] for r in responses}

    stores = {}
    for uid in assigned_ids:
        u = users_by_id.get(uid)
        suc = ((u or {}).get('sucursal') or '').strip() or 'Sin sucursal'
        entry = stores.setdefault(suc, {'sucursal': suc, 'asignados': set(), 'respondieron': set()})
        entry['asignados'].add(uid)
        if uid in responded_ids:
            entry['respondieron'].add(uid)

    por_sucursal = []
    for suc, info in stores.items():
        suc_resp = [r for r in responses if r['respondent_id'] in info['respondieron']]
        yes = no = na = 0
        for r in suc_resp:
            for e in r.get('entries', []):
                v = e.get('respuesta')
                if v == 'Sí':
                    yes += 1
                elif v == 'No':
                    no += 1
                elif v == 'No aplica':
                    na += 1
        evaluated = yes + no
        n_asig, n_resp = len(info['asignados']), len(info['respondieron'])
        estado = 'Completo' if n_resp and n_resp >= n_asig else ('Parcial' if n_resp else 'Pendiente')
        por_sucursal.append({
            'sucursal': suc, 'estado': estado, 'asignados': n_asig, 'respondieron': n_resp,
            'promociones_evaluadas': evaluated,
            'cumplimiento': round(yes / evaluated * 100) if evaluated else None,
        })
    por_sucursal.sort(key=lambda x: x['sucursal'])

    por_promo = {it['id']: {'id': it['id'], 'titulo': it['titulo'], 'visibles': 0, 'no_visibles': 0, 'no_aplica': 0}
                 for it in form.get('items', [])}
    for r in responses:
        for e in r.get('entries', []):
            row = por_promo.get(e['id'])
            if not row:
                continue
            v = e.get('respuesta')
            if v == 'Sí':
                row['visibles'] += 1
            elif v == 'No':
                row['no_visibles'] += 1
            elif v == 'No aplica':
                row['no_aplica'] += 1
    por_promocion = []
    for row in por_promo.values():
        evaluated = row['visibles'] + row['no_visibles']
        row['cumplimiento'] = round(row['visibles'] / evaluated * 100) if evaluated else None
        por_promocion.append(row)
    por_promocion.sort(key=lambda r: r['titulo'])

    total_yes = sum(r['visibles'] for r in por_promocion)
    total_no = sum(r['no_visibles'] for r in por_promocion)
    total_eval = total_yes + total_no
    kpis = {
        'cumplimiento_general': round(total_yes / total_eval * 100) if total_eval else None,
        'tiendas_reportadas': sum(1 for s in por_sucursal if s['respondieron'] > 0),
        'tiendas_total': len(por_sucursal),
        'promociones_evaluadas': len(por_promocion),
        'asignados': len(assigned_ids),
        'respondieron': len(responded_ids),
        'pendientes': max(0, len(assigned_ids) - len(responded_ids)),
    }
    return {'kpis': kpis, 'por_promocion': por_promocion, 'por_sucursal': por_sucursal}


@router.get('/{form_id}/respuestas/{resp_id}')
async def get_response(form_id: str, resp_id: str, user=Depends(get_current_user)):
    form = await _get_form_or_404(form_id)
    row = await db.custom_form_responses.find_one({'id': resp_id, 'form_id': form_id}, {'_id': 0})
    if not row:
        raise HTTPException(status_code=404, detail='Respuesta no encontrada')
    if not (_sees_all_responses(user, form) or row['respondent_id'] == user['id']):
        raise HTTPException(status_code=403, detail='No tienes acceso a esta respuesta')
    return serialize_doc(row)


@router.get('/{form_id}/respuestas/{resp_id}/foto/{photo_id}')
async def get_response_photo(form_id: str, resp_id: str, photo_id: str, user=Depends(get_current_user)):
    form = await _get_form_or_404(form_id)
    row = await db.custom_form_responses.find_one({'id': resp_id, 'form_id': form_id}, {'_id': 0})
    if not row:
        raise HTTPException(status_code=404, detail='Respuesta no encontrada')
    if not (_sees_all_responses(user, form) or row['respondent_id'] == user['id']):
        raise HTTPException(status_code=403, detail='No tienes acceso a esta respuesta')
    photo = None
    for e in row.get('entries', []):
        for p in e.get('photos', []):
            if p['id'] == photo_id:
                photo = p
    if not photo:
        raise HTTPException(status_code=404, detail='Foto no encontrada')
    try:
        content, ctype = await storage.get_object(photo['path'])
    except Exception as ex:
        logger.error(f'photo download failed: {ex}')
        raise HTTPException(status_code=502, detail='No se pudo descargar la foto')
    return Response(content=content, media_type=photo.get('content_type', ctype))
