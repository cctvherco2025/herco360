"""Reportes CAMS — ingest and reporting for the in-store people-counting agent.

The local agent (Proyecto CAM) POSTs batches of "entrada" events to
`/api/cams/ingesta?k=<CAM_INGEST_KEY>`. It retries on any non-2xx and sends a
deterministic `batch_id`, so we deduplicate by that id and always answer 2xx once
a batch is stored.

Environment:
    CAM_INGEST_KEY      shared secret required on the ingest URL (?k=...)
    CAM_PILOT_SUCURSAL   store label for the pilot camera (default "Piloto")
    CAM_PILOT_CAMARA     camera label (default "Camara 1")

Collections:
    cam_lotes    one doc per accepted POST  (batch_id unique)
    cam_eventos  one doc per entry timestamp (date_local / hour_local precomputed)
    cam_agentes  last-seen status per (sucursal, camara)

Event timestamps arrive in UTC ("...Z"); `date_local` / `hour_local` are in the
company timezone (core.APP_UTC_OFFSET_HOURS) so day/hour reports are simple.
"""
import os
import io
import logging
from datetime import datetime, timedelta

from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse

from core import db, get_current_user, serialize_doc, new_id, now_iso, now_local, APP_UTC_OFFSET_HOURS
from models import CamsIngestInput

router = APIRouter(prefix='/cams', tags=['cams'])
logger = logging.getLogger('cams')

INGEST_KEY = os.environ.get('CAM_INGEST_KEY', '').strip()
PILOT_SUCURSAL = os.environ.get('CAM_PILOT_SUCURSAL', 'Piloto').strip()
PILOT_CAMARA = os.environ.get('CAM_PILOT_CAMARA', 'Camara 1').strip()

_OFFSET = timedelta(hours=APP_UTC_OFFSET_HOURS)
ONLINE_WINDOW_MIN = 15  # agent is "en línea" if it reported within this many minutes


def _to_local(ts_utc: str) -> datetime:
    """Parse an ISO-8601 UTC timestamp and shift it to company local time (naive)."""
    s = (ts_utc or '').strip().replace('Z', '').replace('+00:00', '')
    return datetime.fromisoformat(s) + _OFFSET


# --------------------------------------------------------------------------- #
#  Ingest (no user auth — shared key on the URL)
# --------------------------------------------------------------------------- #
@router.post('/ingesta')
async def ingesta(data: CamsIngestInput, k: str = Query(default='')):
    if not INGEST_KEY:
        raise HTTPException(status_code=503, detail='Ingesta de cámaras no configurada en el servidor')
    if k != INGEST_KEY:
        raise HTTPException(status_code=401, detail='Clave inválida')

    sucursal, camara = PILOT_SUCURSAL, PILOT_CAMARA

    # Deduplicate: the agent resends the same batch_id if a response was lost.
    if await db.cam_lotes.find_one({'batch_id': data.batch_id}, {'_id': 1}):
        return {'ok': True, 'dedup': True, 'stored': 0}

    eventos = []
    for ts in (data.eventos or []):
        try:
            loc = _to_local(ts)
        except Exception:
            continue
        eventos.append({
            'id': new_id(),
            'batch_id': data.batch_id,
            'sucursal': sucursal,
            'camara': camara,
            'ts_utc': ts,
            'date_local': loc.strftime('%Y-%m-%d'),
            'hour_local': loc.hour,
            'created_at': now_iso(),
        })

    await db.cam_lotes.insert_one({
        'id': new_id(),
        'batch_id': data.batch_id,
        'sucursal': sucursal,
        'camara': camara,
        'batch_timestamp': data.timestamp,
        'entradas': int(data.entradas or 0),
        'eventos_guardados': len(eventos),
        'received_at': now_iso(),
    })
    if eventos:
        await db.cam_eventos.insert_many(eventos)

    await db.cam_agentes.update_one(
        {'sucursal': sucursal, 'camara': camara},
        {'$set': {'last_seen': now_iso(), 'last_batch_id': data.batch_id},
         '$inc': {'total_entradas': int(data.entradas or 0), 'total_lotes': 1},
         '$setOnInsert': {'id': new_id(), 'first_seen': now_iso()}},
        upsert=True,
    )
    return {'ok': True, 'dedup': False, 'stored': len(eventos)}


# --------------------------------------------------------------------------- #
#  Reporting (admins + Director comercial)
# --------------------------------------------------------------------------- #
async def require_cams_access(user=Depends(get_current_user)):
    if user.get('role') == 'admin' or (user.get('position') or '').strip() == 'Director comercial':
        return user
    raise HTTPException(status_code=403, detail='No tienes acceso al módulo de Reportes CAMS')


async def _agent_state(sucursal=None):
    q = {} if not sucursal else {'sucursal': sucursal}
    agents = await db.cam_agentes.find(q, {'_id': 0}).to_list(50)
    out = []
    now = now_local()
    for a in agents:
        online = False
        mins = None
        try:
            last = _to_local(a['last_seen'])
            mins = (now - last).total_seconds() / 60.0
            online = mins < ONLINE_WINDOW_MIN
        except Exception:
            pass
        out.append({**a, 'online': online, 'mins_since': None if mins is None else round(mins, 1)})
    return out


@router.get('/meta')
async def meta(user=Depends(require_cams_access)):
    sucursales = await db.cam_eventos.distinct('sucursal')
    return {
        'configured': bool(INGEST_KEY),
        'sucursales': sorted([s for s in sucursales if s]),
        'agents': serialize_doc(await _agent_state()),
        'online_window_min': ONLINE_WINDOW_MIN,
    }


@router.get('/summary')
async def summary(start: str, end: str, sucursal: str = None, user=Depends(require_cams_access)):
    q = {'date_local': {'$gte': start, '$lte': end}}
    if sucursal:
        q['sucursal'] = sucursal
    rows = await db.cam_eventos.find(q, {'_id': 0, 'date_local': 1}).to_list(200000)
    by_day = {}
    for r in rows:
        by_day[r['date_local']] = by_day.get(r['date_local'], 0) + 1

    # Fill every day in the range so the chart has no gaps.
    days = []
    try:
        d0 = datetime.strptime(start, '%Y-%m-%d').date()
        d1 = datetime.strptime(end, '%Y-%m-%d').date()
        cur = d0
        while cur <= d1:
            key = cur.isoformat()
            days.append({'date': key, 'count': by_day.get(key, 0)})
            cur += timedelta(days=1)
    except Exception:
        days = [{'date': k, 'count': v} for k, v in sorted(by_day.items())]

    today_key = now_local().strftime('%Y-%m-%d')
    tq = {'date_local': today_key}
    if sucursal:
        tq['sucursal'] = sucursal
    today = await db.cam_eventos.count_documents(tq)

    return {'total': len(rows), 'today': today, 'by_day': days}


@router.get('/hourly')
async def hourly(date: str, sucursal: str = None, user=Depends(require_cams_access)):
    q = {'date_local': date}
    if sucursal:
        q['sucursal'] = sucursal
    rows = await db.cam_eventos.find(q, {'_id': 0, 'hour_local': 1}).to_list(200000)
    hours = [0] * 24
    for r in rows:
        h = r.get('hour_local')
        if isinstance(h, int) and 0 <= h < 24:
            hours[h] += 1
    return {'date': date, 'hours': hours, 'total': len(rows)}


@router.get('/recent')
async def recent(limit: int = 20, sucursal: str = None, user=Depends(require_cams_access)):
    q = {} if not sucursal else {'sucursal': sucursal}
    rows = await db.cam_eventos.find(q, {'_id': 0}).sort('ts_utc', -1).limit(min(limit, 100)).to_list(100)
    for r in rows:
        try:
            r['hora_local'] = _to_local(r['ts_utc']).strftime('%d/%m %H:%M:%S')
        except Exception:
            r['hora_local'] = r.get('ts_utc', '')
    return serialize_doc(rows)


@router.get('/export')
async def export(start: str, end: str, sucursal: str = None, user=Depends(require_cams_access)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    q = {'date_local': {'$gte': start, '$lte': end}}
    if sucursal:
        q['sucursal'] = sucursal
    rows = await db.cam_eventos.find(q, {'_id': 0}).sort('ts_utc', 1).to_list(200000)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Entradas'
    ws.append(['Fecha', 'Hora', 'Sucursal', 'Cámara'])
    head = PatternFill('solid', fgColor='1E395E')
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = head
    for r in rows:
        try:
            loc = _to_local(r['ts_utc'])
            f, h = loc.strftime('%Y-%m-%d'), loc.strftime('%H:%M:%S')
        except Exception:
            f, h = r.get('date_local', ''), ''
        ws.append([f, h, r.get('sucursal', ''), r.get('camara', '')])
    for col, w in zip('ABCD', (14, 12, 12, 22)):
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet('Por día')
    ws2.append(['Fecha', 'Entradas'])
    for c in ws2[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = head
    by_day = {}
    for r in rows:
        by_day[r['date_local']] = by_day.get(r['date_local'], 0) + 1
    for k in sorted(by_day):
        ws2.append([k, by_day[k]])
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    suf = f'_{sucursal}' if sucursal else ''
    fname = f'reportes_cams{suf}_{start}_a_{end}.xlsx'
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )
