"""Inventory module routes (Productos / Tiendas / Movimientos). Restricted access."""
import os
import re
import io
import logging
from datetime import datetime
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from core import db, require_inventory_access, serialize_doc, new_id, now_iso
from models import InventoryIntake, InventoryMovementInput, CatalogItemInput, SUCURSALES

router = APIRouter(prefix='/inventory', tags=['inventory'])


@router.get('/meta')
async def meta(user=Depends(require_inventory_access)):
    return {'sucursales': SUCURSALES}


# ---- Article catalog (autocomplete source) ----
@router.get('/catalog')
async def catalog(q: str = '', user=Depends(require_inventory_access)):
    query = {}
    if q and q.strip():
        query['name_key'] = {'$regex': re.escape(q.strip().lower())}
    items = await db.inventory_catalog.find(query, {'_id': 0, 'name': 1}).sort('name', 1).limit(15).to_list(15)
    return [i['name'] for i in items]


@router.post('/catalog')
async def add_catalog_item(data: CatalogItemInput, user=Depends(require_inventory_access)):
    name = data.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail='Nombre vacío')
    await db.inventory_catalog.update_one(
        {'name_key': name.lower()},
        {'$setOnInsert': {'id': new_id(), 'name': name, 'name_key': name.lower(), 'created_at': now_iso()}},
        upsert=True)
    return {'message': 'ok', 'name': name}


# ---- Article image (uploaded via Excel, stored in DB) ----
logger = logging.getLogger("inventory")


@router.get('/image')
async def article_image(article: str = '', user=Depends(require_inventory_access)):
    """Return the stored product image for an article as a data URL (uploaded via Excel)."""
    name = (article or '').strip()
    if not name:
        raise HTTPException(status_code=400, detail='Indica el artículo')
    key = name.lower()
    cached = await db.inventory_images.find_one({'name_key': key}, {'_id': 0, 'data': 1})
    if cached and cached.get('data'):
        return {'article': name, 'data': cached['data']}
    return {'article': name, 'data': None}


# ---- Stock ----
@router.get('/stock')
async def stock(sucursal: str = None, q: str = None, user=Depends(require_inventory_access)):
    query = {'quantity': {'$gt': 0}}
    if sucursal:
        query['sucursal'] = sucursal
    if q and q.strip():
        query['article_key'] = {'$regex': re.escape(q.strip().lower())}
    items = await db.inventory_stock.find(query, {'_id': 0}).sort('article', 1).to_list(2000)
    return serialize_doc(items)


@router.get('/summary')
async def summary(user=Depends(require_inventory_access)):
    result = []
    for suc in SUCURSALES:
        rows = await db.inventory_stock.find({'sucursal': suc, 'quantity': {'$gt': 0}}, {'_id': 0, 'quantity': 1}).to_list(5000)
        result.append({
            'sucursal': suc,
            'items_count': len(rows),
            'total_qty': sum(r['quantity'] for r in rows),
        })
    return result


# ---- Intake (inventory an article) ----
@router.post('/intake')
async def intake(data: InventoryIntake, user=Depends(require_inventory_access)):
    article = (data.article or '').strip()
    if not article:
        raise HTTPException(status_code=400, detail='Indica el artículo')
    if data.sucursal not in SUCURSALES:
        raise HTTPException(status_code=400, detail='Sucursal inválida')
    if data.quantity is None or data.quantity <= 0:
        raise HTTPException(status_code=400, detail='La cantidad debe ser mayor a 0')

    key = article.lower()
    existing = await db.inventory_stock.find_one({'article_key': key, 'sucursal': data.sucursal})
    if existing:
        new_qty = existing['quantity'] + data.quantity
        await db.inventory_stock.update_one({'id': existing['id']}, {'$set': {'quantity': new_qty, 'article': article, 'updated_at': now_iso()}})
    else:
        new_qty = data.quantity
        await db.inventory_stock.insert_one({
            'id': new_id(), 'article': article, 'article_key': key,
            'sucursal': data.sucursal, 'quantity': data.quantity,
            'created_at': now_iso(), 'updated_at': now_iso(),
        })

    # keep catalog updated for autocomplete
    await db.inventory_catalog.update_one(
        {'name_key': key},
        {'$setOnInsert': {'id': new_id(), 'name': article, 'name_key': key, 'created_at': now_iso()}},
        upsert=True)

    await db.inventory_movements.insert_one({
        'id': new_id(), 'type': 'entrada', 'article': article, 'sucursal': data.sucursal,
        'quantity': data.quantity, 'description': 'Ingreso a inventario', 'solicitante': '',
        'registered_by': user['id'], 'registered_by_name': user['name'],
        'registered_by_avatar': user.get('avatar_url'), 'created_at': now_iso(),
    })
    return {'message': 'Artículo inventariado', 'article': article, 'sucursal': data.sucursal, 'stock': new_qty}


# ---- Movement (rebaja / salida) ----
@router.post('/movement')
async def movement(data: InventoryMovementInput, user=Depends(require_inventory_access)):
    article = (data.article or '').strip()
    if not article:
        raise HTTPException(status_code=400, detail='Indica el artículo')
    if data.sucursal not in SUCURSALES:
        raise HTTPException(status_code=400, detail='Sucursal inválida')
    if data.quantity is None or data.quantity <= 0:
        raise HTTPException(status_code=400, detail='La cantidad debe ser mayor a 0')
    if not (data.description or '').strip():
        raise HTTPException(status_code=400, detail='La descripción del movimiento es obligatoria')

    key = article.lower()
    existing = await db.inventory_stock.find_one({'article_key': key, 'sucursal': data.sucursal})
    available = existing['quantity'] if existing else 0
    if data.quantity > available:
        raise HTTPException(status_code=409, detail=f'Stock insuficiente en {data.sucursal}. Disponible: {available}')

    new_qty = available - data.quantity
    await db.inventory_stock.update_one({'id': existing['id']}, {'$set': {'quantity': new_qty, 'updated_at': now_iso()}})

    await db.inventory_movements.insert_one({
        'id': new_id(), 'type': 'salida', 'article': article, 'sucursal': data.sucursal,
        'quantity': data.quantity, 'description': data.description.strip(),
        'solicitante': (data.solicitante or '').strip(),
        'registered_by': user['id'], 'registered_by_name': user['name'],
        'registered_by_avatar': user.get('avatar_url'), 'created_at': now_iso(),
    })
    return {'message': 'Movimiento registrado', 'article': article, 'sucursal': data.sucursal, 'stock': new_qty}


@router.get('/movements')
async def movements(sucursal: str = None, type: str = None, user=Depends(require_inventory_access)):
    query = {}
    if sucursal:
        query['sucursal'] = sucursal
    if type:
        query['type'] = type
    items = await db.inventory_movements.find(query, {'_id': 0}).sort('created_at', -1).limit(300).to_list(300)
    return serialize_doc(items)



# ====================== REPORTS (Excel / PDF) ======================
async def _get_stock_rows(sucursal=None):
    q = {'quantity': {'$gt': 0}}
    if sucursal:
        q['sucursal'] = sucursal
    return await db.inventory_stock.find(q, {'_id': 0}).sort([('sucursal', 1), ('article', 1)]).to_list(5000)


async def _get_movement_rows(sucursal=None):
    q = {}
    if sucursal:
        q['sucursal'] = sucursal
    return await db.inventory_movements.find(q, {'_id': 0}).sort('created_at', -1).limit(2000).to_list(2000)


def _ts():
    return datetime.now().strftime('%Y%m%d_%H%M')


def _build_stock_xlsx(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    ws.append(['Artículo', 'Sucursal', 'Cantidad'])
    head_fill = PatternFill('solid', fgColor='1E395E')
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = head_fill
        c.alignment = Alignment(horizontal='center')
    for r in rows:
        ws.append([r.get('article', ''), r.get('sucursal', ''), r.get('quantity', 0)])
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws2 = wb.create_sheet('Resumen')
    ws2.append(['Sucursal', 'Artículos', 'Unidades'])
    for c in ws2[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = head_fill
    for suc in SUCURSALES:
        srows = [r for r in rows if r.get('sucursal') == suc]
        ws2.append([suc, len(srows), sum(r.get('quantity', 0) for r in srows)])
    for col in ('A', 'B', 'C'):
        ws2.column_dimensions[col].width = 14
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def _build_movements_xlsx(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = 'Movimientos'
    ws.append(['Fecha', 'Tipo', 'Artículo', 'Sucursal', 'Cantidad', 'Descripción', 'Solicitante', 'Registrado por'])
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='1E395E')
    for r in rows:
        ws.append([
            (r.get('created_at') or '')[:19].replace('T', ' '),
            'Salida' if r.get('type') == 'salida' else 'Entrada',
            r.get('article', ''), r.get('sucursal', ''),
            (-r.get('quantity', 0) if r.get('type') == 'salida' else r.get('quantity', 0)),
            r.get('description', ''), r.get('solicitante', ''), r.get('registered_by_name', ''),
        ])
    widths = [20, 10, 50, 12, 12, 40, 22, 20]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def _build_pdf(title, columns, data_rows, col_widths):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    elems = [Paragraph(f'<b>{title}</b>', styles['Title']),
             Paragraph(datetime.now().strftime('Generado: %d/%m/%Y %H:%M'), styles['Normal']),
             Spacer(1, 8)]
    cell = styles['BodyText']; cell.fontSize = 8; cell.leading = 10
    table_data = [columns] + [[Paragraph(str(c), cell) for c in row] for row in data_rows]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E395E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F5FA')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D8DEE9')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elems.append(t)
    doc.build(elems)
    buf.seek(0)
    return buf


def _stream(buf, media_type, filename):
    return StreamingResponse(buf, media_type=media_type,
                             headers={'Content-Disposition': f'attachment; filename="{filename}"'})


@router.get('/export/stock')
async def export_stock(format: str = 'xlsx', sucursal: str = None, user=Depends(require_inventory_access)):
    rows = await _get_stock_rows(sucursal)
    suffix = f'_{sucursal}' if sucursal else ''
    if format == 'pdf':
        from reportlab.lib.units import mm
        data = [[r.get('article', ''), r.get('sucursal', ''), r.get('quantity', 0)] for r in rows]
        buf = _build_pdf('HERCO360 - Reporte de Inventario', ['Artículo', 'Sucursal', 'Cantidad'],
                         data, [200 * mm, 30 * mm, 30 * mm])
        return _stream(buf, 'application/pdf', f'inventario{suffix}_{_ts()}.pdf')
    buf = _build_stock_xlsx(rows)
    return _stream(buf, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                   f'inventario{suffix}_{_ts()}.xlsx')


@router.get('/export/movements')
async def export_movements(format: str = 'xlsx', sucursal: str = None, user=Depends(require_inventory_access)):
    rows = await _get_movement_rows(sucursal)
    if format == 'pdf':
        from reportlab.lib.units import mm
        data = [[(r.get('created_at') or '')[:16].replace('T', ' '),
                 'Salida' if r.get('type') == 'salida' else 'Entrada',
                 r.get('article', ''), r.get('sucursal', ''),
                 (-r.get('quantity', 0) if r.get('type') == 'salida' else r.get('quantity', 0)),
                 r.get('description', ''), r.get('solicitante', '')] for r in rows]
        buf = _build_pdf('HERCO360 - Reporte de Movimientos',
                         ['Fecha', 'Tipo', 'Artículo', 'Sucursal', 'Cant.', 'Descripción', 'Solicitante'],
                         data, [28 * mm, 18 * mm, 90 * mm, 20 * mm, 16 * mm, 60 * mm, 35 * mm])
        return _stream(buf, 'application/pdf', f'movimientos_{_ts()}.pdf')
    buf = _build_movements_xlsx(rows)
    return _stream(buf, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                   f'movimientos_{_ts()}.xlsx')


# ====================== IMPORT (bulk articles via Excel) ======================
@router.get('/template')
async def import_template(user=Depends(require_inventory_access)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = 'Articulos'
    ws.append(['Articulo', 'Cantidad', 'Sucursal'])
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='00A5DF')
        c.alignment = Alignment(horizontal='center')
    ws.append(['Dual hook', 100, 'H1'])
    ws.append(['Wire basket, 1000mm*470*250mm', 25, 'H2'])
    ws.append(['Nuevo artículo de ejemplo', '', ''])
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws2 = wb.create_sheet('Instrucciones')
    for row in [
        ['Instrucciones para importar artículos'],
        [''],
        ['Columna', 'Descripción'],
        ['Articulo', 'Obligatorio. Nombre del artículo. Si no existe, se agrega al catálogo.'],
        ['Cantidad', 'Opcional. Si la indicas (número > 0), se suma al inventario.'],
        ['Sucursal', 'Opcional. Requerida solo si pones Cantidad. Valores: ' + ', '.join(SUCURSALES)],
        [''],
        ['Nota', 'Si solo pones Articulo, se agrega al catálogo (sin stock).'],
    ]:
        ws2.append(row)
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 80
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return _stream(buf, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                   'plantilla_articulos_herco360.xlsx')


@router.post('/import')
async def import_articles(file: UploadFile = File(...), user=Depends(require_inventory_access)):
    fn = (file.filename or '').lower()
    if not fn.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail='El archivo debe ser .xlsx')
    from openpyxl import load_workbook
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail='No se pudo leer el archivo Excel')
    ws = wb['Articulos'] if 'Articulos' in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail='El archivo está vacío')

    header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]

    def col_idx(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    i_art = col_idx('articulo', 'artículo', 'article')
    i_qty = col_idx('cantidad', 'quantity', 'cant')
    i_suc = col_idx('sucursal', 'branch')
    start = 1 if i_art is not None else 0
    if i_art is None:
        i_art = 0

    added_catalog = 0
    stock_entries = 0
    errors = []
    for ln, row in enumerate(rows[start:], start=start + 1):
        if not row:
            continue
        article = row[i_art] if i_art < len(row) else None
        if article is None or not str(article).strip():
            continue
        article = str(article).strip()
        key = article.lower()
        res = await db.inventory_catalog.update_one(
            {'name_key': key},
            {'$setOnInsert': {'id': new_id(), 'name': article, 'name_key': key, 'created_at': now_iso()}},
            upsert=True)
        if res.upserted_id is not None:
            added_catalog += 1
        qty = row[i_qty] if (i_qty is not None and i_qty < len(row)) else None
        suc = row[i_suc] if (i_suc is not None and i_suc < len(row)) else None
        if qty not in (None, ''):
            try:
                qty = int(float(qty))
            except Exception:
                errors.append(f'Fila {ln}: cantidad inválida'); continue
            if qty <= 0:
                continue
            suc = str(suc).strip().upper() if suc else None
            if suc not in SUCURSALES:
                errors.append(f'Fila {ln}: sucursal inválida ("{suc}")'); continue
            existing = await db.inventory_stock.find_one({'article_key': key, 'sucursal': suc})
            if existing:
                await db.inventory_stock.update_one({'id': existing['id']}, {'$set': {'quantity': existing['quantity'] + qty, 'article': article, 'updated_at': now_iso()}})
            else:
                await db.inventory_stock.insert_one({'id': new_id(), 'article': article, 'article_key': key, 'sucursal': suc, 'quantity': qty, 'created_at': now_iso(), 'updated_at': now_iso()})
            await db.inventory_movements.insert_one({
                'id': new_id(), 'type': 'entrada', 'article': article, 'sucursal': suc, 'quantity': qty,
                'description': 'Importación desde Excel', 'solicitante': '',
                'registered_by': user['id'], 'registered_by_name': user['name'],
                'registered_by_avatar': user.get('avatar_url'), 'created_at': now_iso(),
            })
            stock_entries += 1
    wb.close()
    return {
        'message': 'Importación completada',
        'added_to_catalog': added_catalog,
        'stock_entries': stock_entries,
        'errors': errors[:20],
    }



# ====================== IMPORT IMAGES (Excel with embedded pictures) ======================
# Namespaces used inside an .xlsx archive for drawings / images.
_NS_XDR = '{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}'
_NS_A = '{http://schemas.openxmlformats.org/drawingml/2006/main}'
_NS_R = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'
_NS_REL = '{http://schemas.openxmlformats.org/package/2006/relationships}'

_MIME_BY_EXT = {
    'png': 'image/png', 'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
    'gif': 'image/gif', 'bmp': 'image/bmp', 'webp': 'image/webp',
}


def _parse_xlsx_images(content: bytes):
    """Extract embedded images from an .xlsx file mapped to their anchor row.
    Returns a list of tuples: (row_index_0based, mime_type, image_bytes)."""
    import zipfile
    import posixpath
    import xml.etree.ElementTree as ET

    results = []
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        names = zf.namelist()
        drawings = [n for n in names if re.match(r'xl/drawings/drawing\d+\.xml$', n)]
        for drawing in drawings:
            # Map relationship ids -> media file path for this drawing.
            rels_path = posixpath.join('xl/drawings/_rels', posixpath.basename(drawing) + '.rels')
            rel_map = {}
            if rels_path in names:
                try:
                    rels_root = ET.fromstring(zf.read(rels_path))
                    for rel in rels_root.findall(_NS_REL + 'Relationship'):
                        rid = rel.get('Id')
                        target = rel.get('Target')
                        if rid and target:
                            if target.startswith('/'):
                                # Absolute within the package (relative to archive root).
                                rel_map[rid] = target.lstrip('/')
                            else:
                                # Relative to xl/drawings/
                                rel_map[rid] = posixpath.normpath(posixpath.join('xl/drawings', target))
                except Exception:
                    pass
            try:
                root = ET.fromstring(zf.read(drawing))
            except Exception:
                continue
            for anchor in list(root):
                tag = anchor.tag.split('}')[-1]
                if tag not in ('oneCellAnchor', 'twoCellAnchor', 'absoluteAnchor'):
                    continue
                from_el = anchor.find(_NS_XDR + 'from')
                row_idx = 0
                if from_el is not None:
                    row_el = from_el.find(_NS_XDR + 'row')
                    if row_el is not None and row_el.text is not None:
                        try:
                            row_idx = int(row_el.text)
                        except Exception:
                            row_idx = 0
                blip = anchor.find('.//' + _NS_A + 'blip')
                if blip is None:
                    continue
                embed = blip.get(_NS_R + 'embed')
                media_path = rel_map.get(embed)
                if not media_path or media_path not in names:
                    continue
                ext = media_path.rsplit('.', 1)[-1].lower()
                mime = _MIME_BY_EXT.get(ext, 'image/png')
                try:
                    img_bytes = zf.read(media_path)
                except Exception:
                    continue
                results.append((row_idx, mime, img_bytes))
    return results


@router.get('/images-template')
async def images_template(user=Depends(require_inventory_access)):
    """Downloadable Excel template explaining how to attach an image per article."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = 'Imagenes'
    ws.append(['Articulo', 'Imagen'])
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='00A5DF')
        c.alignment = Alignment(horizontal='center')
    ws.append(['Dual hook', '(pega aquí la imagen)'])
    ws.append(['Wire basket, 1000mm*470*250mm', '(pega aquí la imagen)'])
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 40
    for r in range(2, 30):
        ws.row_dimensions[r].height = 90
    ws2 = wb.create_sheet('Instrucciones')
    for row in [
        ['Cómo subir imágenes de artículos'],
        [''],
        ['1) En la columna A (Articulo) escribe el nombre EXACTO del artículo.'],
        ['2) En la columna B (Imagen), inserta/pega la foto en la MISMA fila del artículo.'],
        ['   (En Excel: Insertar > Imágenes, y colócala sobre la celda de esa fila).'],
        ['3) Guarda el archivo como .xlsx y súbelo en el sistema.'],
        [''],
        ['Notas:'],
        ['- El nombre del artículo debe coincidir con el del catálogo.'],
        ['- Una imagen por fila. Formatos: PNG, JPG.'],
        ['- Si subes una imagen para un artículo que ya tenía, se reemplaza.'],
    ]:
        ws2.append(row)
    ws2.column_dimensions['A'].width = 80
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return _stream(buf, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                   'plantilla_imagenes_herco360.xlsx')


@router.post('/import-images')
async def import_images(file: UploadFile = File(...), user=Depends(require_inventory_access)):
    """Import an Excel file with embedded images and assign each image to its article (by row)."""
    fn = (file.filename or '').lower()
    if not fn.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail='El archivo debe ser .xlsx')
    content = await file.read()

    # 1) Read article names per row from the sheet.
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail='No se pudo leer el archivo Excel')
    ws = wb['Imagenes'] if 'Imagenes' in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise HTTPException(status_code=400, detail='El archivo está vacío')

    header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
    art_col = 0
    for cand in ('articulo', 'artículo', 'article', 'descripcion', 'descripción', 'nombre'):
        if cand in header:
            art_col = header.index(cand)
            break
    # Map 0-based row index -> article name
    row_to_article = {}
    for i, row in enumerate(rows):
        if i == 0 and any(h in header for h in ('articulo', 'artículo', 'article', 'descripcion', 'descripción', 'nombre')):
            continue  # skip header row
        if row and art_col < len(row) and row[art_col] is not None and str(row[art_col]).strip():
            row_to_article[i] = str(row[art_col]).strip()

    # 2) Extract embedded images mapped to their anchor row.
    try:
        images = _parse_xlsx_images(content)
    except Exception as e:
        logger.warning(f'image parse failed: {e}')
        raise HTTPException(status_code=400, detail='No se pudieron leer las imágenes del Excel')

    if not images:
        raise HTTPException(status_code=400, detail='No se encontraron imágenes incrustadas en el Excel. Inserta las fotos dentro del archivo (Insertar > Imágenes).')

    import base64
    saved = 0
    errors = []
    matched_rows = set()
    for row_idx, mime, img_bytes in images:
        # The anchor row may sit on the article row or one below the picture's top edge.
        article = row_to_article.get(row_idx) or row_to_article.get(row_idx + 1) or row_to_article.get(row_idx - 1)
        if not article:
            errors.append(f'Fila {row_idx + 1}: imagen sin artículo asociado')
            continue
        if len(img_bytes) > 5 * 1024 * 1024:
            errors.append(f'{article}: imagen demasiado grande (>5MB)')
            continue
        b64 = base64.b64encode(img_bytes).decode('utf-8')
        data_url = f'data:{mime};base64,{b64}'
        key = article.lower()
        await db.inventory_images.update_one(
            {'name_key': key},
            {'$set': {'name': article, 'name_key': key, 'data': data_url, 'updated_at': now_iso()},
             '$setOnInsert': {'id': new_id(), 'created_at': now_iso()}},
            upsert=True)
        # Keep catalog in sync so the article is searchable.
        await db.inventory_catalog.update_one(
            {'name_key': key},
            {'$setOnInsert': {'id': new_id(), 'name': article, 'name_key': key, 'created_at': now_iso()}},
            upsert=True)
        saved += 1
        matched_rows.add(row_idx)

    return {
        'message': 'Importación de imágenes completada',
        'images_found': len(images),
        'images_saved': saved,
        'errors': errors[:20],
    }
